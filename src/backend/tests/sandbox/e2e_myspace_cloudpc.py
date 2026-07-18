"""End-to-end live test script for the "My Space" cloud PC (real PG + real OSS).

Real opensandbox is unreachable on this machine → inject a FakeSandbox that faithfully
simulates put/get/execute with a real temp directory + subprocess. The DB and object
storage are both real, so folder-tree creation / artifact reverse-sync / lazy loading /
delete-move are all verified for real.

How to run (inside the container): python tests/e2e_myspace_cloudpc.py
All artifacts use the __e2e__ prefix, with hard cleanup at the end (artifact rows + folder rows + OSS objects + cache).
"""

from __future__ import annotations

import asyncio
import json
import os
import shutil
import subprocess
import tempfile
import traceback
import uuid

USER_ID = "user_a9e0c627c2674456"
CHAT_ID = "chat_d6307bb5ed9049f2"
TAG = f"__e2e__{uuid.uuid4().hex[:6]}"

results: list[tuple[str, bool, str]] = []


def check(name: str, cond: bool, detail: str = "") -> None:
    results.append((name, bool(cond), detail))
    print(f"[{'PASS' if cond else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))


# ── FakeSandbox: real temp directory + subprocess, /workspace → tmp rewrite ──────────
class FakeSandbox:
    def __init__(self) -> None:
        self.root = tempfile.mkdtemp(prefix="fakesbx_")

    def _phys(self, path: str) -> str:
        assert path.startswith("/workspace"), path
        return os.path.join(self.root, path.lstrip("/"))

    async def get_file(self, session_id, path: str) -> bytes:
        from core.sandbox import SandboxError

        fp = self._phys(path)
        if not os.path.isfile(fp):
            raise SandboxError(f"not found: {path}")
        with open(fp, "rb") as f:
            return f.read()

    async def put_file(self, session_id, path: str, content: bytes) -> None:
        fp = self._phys(path)
        os.makedirs(os.path.dirname(fp), exist_ok=True)
        with open(fp, "wb") as f:
            f.write(content)

    async def execute(self, req):
        from core.sandbox.protocol import ExecuteResult

        script = req.script_content.replace("/workspace", f"{self.root}/workspace")
        os.makedirs(f"{self.root}/workspace", exist_ok=True)
        proc = subprocess.run(
            ["bash", "-c", script], capture_output=True, text=True,
            timeout=req.timeout or 30,
        )
        return ExecuteResult(
            stdout=proc.stdout, stderr=proc.stderr,
            exit_code=proc.returncode, execution_time_ms=1, files=[],
        )

    async def stage_files(self, user_id, files):
        return []


FAKE = FakeSandbox()


def _install_fake_provider():
    import core.sandbox as sb
    import core.sandbox.factory as fac

    sb.get_sandbox_provider = lambda: FAKE          # type: ignore
    fac.get_sandbox_provider = lambda: FAKE          # type: ignore
    fac._provider = FAKE                              # type: ignore


# ── Use a fake Toolkit to capture each tool's inner function ────────────────────────────────
class CapToolkit:
    def __init__(self) -> None:
        self.fns: dict = {}

    def register_tool_function(self, fn, **kw):
        self.fns[fn.__name__] = fn


def _resp(tool_resp) -> dict:
    """ToolResponse → dict (tools all use a single TextBlock JSON)."""
    blk = tool_resp.content[0]
    txt = blk["text"] if isinstance(blk, dict) else blk.text
    return json.loads(txt)


async def main() -> None:
    _install_fake_provider()

    from core.db.engine import SessionLocal
    from core.db.models import Artifact, UserFolder
    from core.llm.tools import (
        ReadStateTracker,
        register_delete,
        register_edit,
        register_glob,
        register_grep,
        register_move,
        register_read,
        register_write,
    )
    from core.storage import get_storage

    tk = CapToolkit()
    st = ReadStateTracker()
    kw = dict(chat_id=CHAT_ID, user_id=USER_ID)
    register_read(tk, state=st, **kw)
    register_edit(tk, state=st, **kw)
    register_write(tk, state=st, **kw)
    register_delete(tk, state=st, **kw)
    register_move(tk, state=st, **kw)
    register_glob(tk, **kw)
    register_grep(tk, **kw)
    Read, Edit, Write = tk.fns["Read"], tk.fns["Edit"], tk.fns["Write"]
    Delete, Move = tk.fns["Delete"], tk.fns["Move"]
    Glob, Grep = tk.fns["Glob"], tk.fns["Grep"]

    created_storage_keys: list[str] = []

    def track_keys():
        db = SessionLocal()
        try:
            rows = db.query(Artifact.storage_key).filter(
                Artifact.user_id == USER_ID,
                Artifact.filename.like(f"%{TAG}%"),
            ).all()
            for (k,) in rows:
                if k and k not in created_storage_keys:
                    created_storage_keys.append(k)
        finally:
            db.close()

    base = f"/myspace/{TAG}_dir/sub"
    f1 = f"{base}/a_{TAG}.txt"

    try:
        # 1. Write a new file to a nested path → folder chain created + artifact lands user_folder_id
        r = _resp(await Write(file_path=f1, content="hello cloud-pc\nline2\n"))
        check("1.Write 嵌套新文件", r.get("ok") and r.get("persistent"), str(r)[:160])
        fid1 = r.get("file_id")
        track_keys()
        db = SessionLocal()
        try:
            from core.llm.tools.myspace_vfs import resolve_folder_id, resolve_artifact
            fr = resolve_folder_id(db, USER_ID, [f"{TAG}_dir", "sub"], create=False)
            art = resolve_artifact(db, USER_ID, fr.folder_id, f"a_{TAG}.txt") if fr.found else None
            check("1b.folder 链已建+artifact 归位",
                  fr.found and art is not None and art.user_folder_id == fr.folder_id,
                  f"found={fr.found} folder={fr.folder_id} art={getattr(art,'artifact_id',None)}")
        finally:
            db.close()

        # 2. Read it back
        r = _resp(await Read(file_path=f1))
        check("2.Read 读回内容", r.get("type") == "text"
              and "hello cloud-pc" in r.get("content", ""), str(r)[:160])

        # 3. Edit in place → same file_id
        r = _resp(await Read(file_path=f1))  # must fully Read first
        r = _resp(await Edit(file_path=f1, old_string="line2", new_string="LINE2-edited"))
        ok_edit = r.get("ok") or r.get("type") == "update" or "diff" in r
        fid_after = r.get("file_id")
        check("3.Edit 就地修改", ok_edit, str(r)[:160])
        r = _resp(await Read(file_path=f1))
        check("3b.Edit 后内容变更+同file_id",
              "LINE2-edited" in r.get("content", "") and (fid_after in (None, fid1)),
              f"fid1={fid1} after={fid_after}")

        # 4. Lazy load: after deleting the sandbox physical copy, Read should materialize it back from the artifact
        phys = os.path.join(FAKE.root, f"workspace/myspace/{USER_ID}/{TAG}_dir/sub/a_{TAG}.txt")
        os.remove(phys)
        st.forget(f1)
        st.forget(f"/workspace/myspace/{USER_ID}/{TAG}_dir/sub/a_{TAG}.txt")
        r = _resp(await Read(file_path=f1))
        check("4.懒加载物化", r.get("type") == "text"
              and "LINE2-edited" in r.get("content", ""), str(r)[:160])

        # 5. Binary office document parsed-text fallback (xlsx)
        try:
            import openpyxl  # noqa: F401
            from openpyxl import Workbook

            wb = Workbook(); ws = wb.active
            ws["A1"] = "营收"; ws["B1"] = "数值"; ws["A2"] = "Q1"; ws["B2"] = 12345
            import io
            buf = io.BytesIO(); wb.save(buf)
            xlsx_logical = f"/myspace/{TAG}_dir/report_{TAG}.xlsx"
            from core.llm.tools import myspace_vfs as _ms
            ref = _ms.sync_upsert(user_id=USER_ID, chat_id=CHAT_ID,
                                  logical_path=xlsx_logical, content=buf.getvalue())
            track_keys()
            await FAKE.put_file(CHAT_ID,
                f"/workspace/myspace/{USER_ID}/{TAG}_dir/report_{TAG}.xlsx", buf.getvalue())
            r = _resp(await Read(file_path=xlsx_logical))
            check("5.xlsx 解析文本回退",
                  r.get("parsed_text") is True and "12345" in r.get("content", ""),
                  str(r)[:160])
        except ImportError:
            check("5.xlsx 解析文本回退", True, "SKIP: openpyxl 不可用")

        # 6. Glob —— myspace goes through the DB tree
        r = _resp(await Glob(pattern=f"**/a_{TAG}.txt", path="/myspace"))
        hit = any(f1.split("/")[-1] in p for p in r.get("filenames", []))
        check("6.Glob 递归命中(DB树)",
              r.get("ok") and r.get("source") == "myspace_tree" and hit, str(r)[:200])
        r2 = _resp(await Glob(pattern="*.txt", path=base))
        check("6b.Glob 非递归当层",
              any(f"a_{TAG}.txt" in p for p in r2.get("filenames", [])), str(r2)[:160])

        # 7. Grep —— content search after subtree materialization
        r = _resp(await Grep(pattern="LINE2-edited", path="/myspace",
                             output_mode="files_with_matches"))
        check("7.Grep 内容命中", r.get("ok") and r.get("num_matches", 0) >= 1,
              f"materialized={r.get('myspace_materialized')} matches={r.get('num_matches')}")

        # 8. Move file rename + change folder
        dst = f"/myspace/{TAG}_dir/renamed_{TAG}.txt"
        r = _resp(await Move(src_path=f1, dst_path=dst))
        check("8.Move 文件", r.get("ok") and r.get("kind") == "file", str(r)[:160])
        db = SessionLocal()
        try:
            from core.llm.tools.myspace_vfs import resolve_folder_id, resolve_artifact
            fr_old = resolve_folder_id(db, USER_ID, [f"{TAG}_dir", "sub"], create=False)
            old = resolve_artifact(db, USER_ID, fr_old.folder_id, f"a_{TAG}.txt") if fr_old.found else None
            fr_new = resolve_folder_id(db, USER_ID, [f"{TAG}_dir"], create=False)
            new = resolve_artifact(db, USER_ID, fr_new.folder_id, f"renamed_{TAG}.txt")
            check("8b.Move 后旧没了/新在", old is None and new is not None,
                  f"old={old} new={getattr(new,'artifact_id',None)}")
        finally:
            db.close()

        # 9. Move folder rename
        r = _resp(await Move(src_path=f"/myspace/{TAG}_dir",
                             dst_path=f"/myspace/{TAG}_archive"))
        check("9.Move 文件夹", r.get("ok") and r.get("kind") == "folder", str(r)[:160])
        db = SessionLocal()
        try:
            from core.llm.tools.myspace_vfs import resolve_folder_id, resolve_artifact
            fr = resolve_folder_id(db, USER_ID, [f"{TAG}_archive"], create=False)
            moved = resolve_artifact(db, USER_ID, fr.folder_id, f"renamed_{TAG}.txt") if fr.found else None
            check("9b.文件夹改名后子文件仍可达", fr.found and moved is not None,
                  f"found={fr.found}")
        finally:
            db.close()

        # 10. Negative case: writing to scratch does not enter My Space
        r = _resp(await Write(file_path=f"/workspace/scratch/tmp_{TAG}.txt",
                              content="scratch only"))
        check("10.scratch 不同步", r.get("ok") and not r.get("persistent")
              and "file_id" not in r, str(r)[:140])

        # 11. Negative case: deleting the root is rejected
        r = _resp(await Delete(path="/myspace"))
        check("11.删根被拒", "error" in r, str(r)[:120])

        # 12. Negative case: Move target already exists
        r = _resp(await Write(file_path=f"/myspace/{TAG}_archive/exists_{TAG}.txt",
                              content="x")); track_keys()
        r = _resp(await Move(src_path=f"/myspace/{TAG}_archive/renamed_{TAG}.txt",
                             dst_path=f"/myspace/{TAG}_archive/exists_{TAG}.txt"))
        check("12.Move 目标存在被拒", "error" in r, str(r)[:140])

        # 13. Delete file soft-delete
        r = _resp(await Delete(path=f"/myspace/{TAG}_archive/renamed_{TAG}.txt"))
        check("13.Delete 文件", r.get("ok") and r.get("kind") == "file", str(r)[:140])
        db = SessionLocal()
        try:
            from core.llm.tools.myspace_vfs import resolve_folder_id, resolve_artifact
            fr = resolve_folder_id(db, USER_ID, [f"{TAG}_archive"], create=False)
            gone = resolve_artifact(db, USER_ID, fr.folder_id, f"renamed_{TAG}.txt")
            check("13b.软删后解析不到", gone is None, f"gone={gone}")
        finally:
            db.close()

        # 14. Delete folder cascade
        r = _resp(await Delete(path=f"/myspace/{TAG}_archive"))
        check("14.Delete 文件夹级联",
              r.get("ok") and r.get("kind") == "folder"
              and r.get("artifacts_affected", -1) >= 1, str(r)[:160])
        db = SessionLocal()
        try:
            from core.llm.tools.myspace_vfs import resolve_folder_id
            fr = resolve_folder_id(db, USER_ID, [f"{TAG}_archive"], create=False)
            check("14b.文件夹已不可达", not fr.found, f"found={fr.found}")
        finally:
            db.close()

    except Exception:
        print("EXCEPTION during test:\n" + traceback.format_exc())
        results.append(("UNCAUGHT-EXCEPTION", False, "see traceback"))
    finally:
        # ── Hard cleanup: OSS objects + artifact rows + folder rows + cache ──
        track_keys()
        storage = get_storage()
        for k in created_storage_keys:
            try:
                storage.delete(k)
            except Exception as e:
                print(f"  cleanup storage {k}: {e}")
        db = SessionLocal()
        try:
            db.query(Artifact).filter(
                Artifact.user_id == USER_ID,
                Artifact.filename.like(f"%{TAG}%"),
            ).delete(synchronize_session=False)
            db.query(UserFolder).filter(
                UserFolder.user_id == USER_ID,
                UserFolder.name.like(f"%{TAG}%"),
            ).delete(synchronize_session=False)
            db.commit()
        finally:
            db.close()
        try:
            from core.sandbox._common import myspace_cache_dir
            cdir = myspace_cache_dir(USER_ID)
            for p in list(cdir.glob(f"**/*{TAG}*")):
                if p.is_file():
                    p.unlink()
        except Exception:
            pass
        shutil.rmtree(FAKE.root, ignore_errors=True)

        n_pass = sum(1 for _, ok, _ in results if ok)
        n_total = len(results)
        print(f"\n==== SUMMARY: {n_pass}/{n_total} PASS ====")
        for name, ok, det in results:
            if not ok:
                print(f"  FAIL: {name} — {det}")
        print("CLEANUP DONE (tag=%s)" % TAG)
        raise SystemExit(0 if n_pass == n_total else 1)


if __name__ == "__main__":
    asyncio.run(main())
