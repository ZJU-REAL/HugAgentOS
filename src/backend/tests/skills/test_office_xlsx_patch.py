"""excel-editing skill engine — model_builder / patch_editor / chart_builder.

Ported from the minimax-xlsx skill. The engine (create/edit) is pure stdlib;
chart_builder needs openpyxl.
"""
import importlib.util
import tempfile
import zipfile
from pathlib import Path

import pytest

from tests._skill_engine import load_engine

load_engine("excel-editing", "xlsx_engine")


_MODEL = {
    "sheets": [
        {"name": "假设", "freeze_header": True,
         "columns": [{"width": 20}, {"width": 14}],
         "rows": [
             {"role": "header", "cells": ["项目", "值"]},
             {"cells": [{"value": "单价", "role": "default"},
                        {"value": 100, "role": "input"}]},
             {"cells": [{"value": "数量", "role": "default"},
                        {"value": 12, "role": "input_int"}]},
             {"cells": [{"value": "小计", "role": "default"},
                        {"formula": "B2*B3", "role": "formula"}]},
         ]},
        {"name": "汇总",
         "rows": [
             {"role": "header", "cells": ["合计"]},
             {"cells": [{"formula": "'假设'!B4", "role": "xref"}]},
         ]},
    ]
}


def test_model_builder_formula_first():
    from xlsx_engine._handle import use_workdir
    from xlsx_engine import model_builder

    with tempfile.TemporaryDirectory() as tmp:
        with use_workdir(tmp):
            res = model_builder.build_model(
                spec=_MODEL, output_filename="model.xlsx")
        mp = Path(tmp) / "model.xlsx"
        assert mp.is_file() and mp.stat().st_size > 0
        assert res["sheets"] == ["假设", "汇总"]
        with zipfile.ZipFile(mp) as z:
            sheet1 = z.read("xl/worksheets/sheet1.xml").decode()
        # formula-first: emitted as a live <f>, never a hardcoded value
        assert "<f>B2*B3</f>" in sheet1


def test_patch_editor_byte_preserving_ops():
    from xlsx_engine._handle import use_workdir
    from xlsx_engine import model_builder, patch_editor

    with tempfile.TemporaryDirectory() as tmp:
        with use_workdir(tmp):
            model_builder.build_model(spec=_MODEL, output_filename="m.xlsx")
            res = patch_editor.apply_patches(
                input_filename="m.xlsx",
                output_filename="m2.xlsx",
                patches=[
                    {"op": "set_cell", "sheet": "假设", "cell": "B2",
                     "value": 150, "role": "input"},
                    {"op": "insert_row", "sheet": "假设", "at": 5,
                     "text": {"A": "税率"}, "values": {"B": 0.06}},
                    {"op": "rename_sheet", "from": "汇总", "to": "Summary"},
                ],
            )
        out = Path(tmp) / "m2.xlsx"
        assert out.is_file()
        assert res["patches_applied"] == 3
        with zipfile.ZipFile(out) as z:
            wbxml = z.read("xl/workbook.xml").decode()
        assert "Summary" in wbxml and "汇总" not in wbxml


def test_patch_editor_rejects_empty():
    from xlsx_engine import patch_editor

    with pytest.raises(patch_editor.XlsxPatchError):
        patch_editor.apply_patches(
            input_filename="x.xlsx", output_filename="y.xlsx", patches=[])


@pytest.mark.skipif(
    importlib.util.find_spec("openpyxl") is None,
    reason="openpyxl only installed in the mcp container",
)
def test_patch_editor_registers_sharedstrings_on_first_string_intro():
    """Regression: when a source .xlsx has no xl/sharedStrings.xml and a
    patch introduces the first string cell (add_column header / set_cell
    str value), the engine must register the new part in BOTH
    [Content_Types].xml and xl/_rels/workbook.xml.rels — otherwise Excel /
    LibreOffice silently render string cells as blank and openpyxl raises
    IndexError trying to look up the missing shared-strings table.
    """
    from xlsx_engine._handle import use_workdir
    from xlsx_engine import builder, patch_editor

    with tempfile.TemporaryDirectory() as tmp:
        with use_workdir(tmp):
            # build a workbook that contains ONLY numeric cells — no string
            # body, so the resulting archive has no xl/sharedStrings.xml
            builder.create_workbook(
                filename="numeric.xlsx",
                sheets=[{"name": "S", "rows": [[1, 2, 3], [4, 5, 6]]}],
            )
            src = Path(tmp) / "numeric.xlsx"
            with zipfile.ZipFile(src) as z:
                assert "xl/sharedStrings.xml" not in z.namelist(), (
                    "fixture assumption broken — workbook unexpectedly contains "
                    "a shared-strings part"
                )

            # add a string-bearing column header AND a string set_cell op
            patch_editor.apply_patches(
                input_filename="numeric.xlsx",
                output_filename="patched.xlsx",
                patches=[
                    {"op": "add_column", "sheet": "S", "col": "D", "header": "毛利率"},
                    {"op": "set_cell", "sheet": "S", "cell": "E1", "value": "备注"},
                ],
            )

        out = Path(tmp) / "patched.xlsx"
        assert out.is_file()
        with zipfile.ZipFile(out) as z:
            assert "xl/sharedStrings.xml" in z.namelist()
            ct = z.read("[Content_Types].xml").decode("utf-8")
            rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        # Manifest must declare the new part — this is the bug surface.
        assert "sharedStrings.xml" in ct, (
            f"[Content_Types].xml is missing sharedStrings Override:\n{ct}"
        )
        assert "sharedStrings.xml" in rels, (
            f"workbook.xml.rels is missing sharedStrings Relationship:\n{rels}"
        )

        # The harder check — openpyxl must read the headers back without
        # blowing up. Pre-fix this raised IndexError.
        from openpyxl import load_workbook

        wb = load_workbook(str(out))
        ws = wb["S"]
        assert ws.cell(row=1, column=4).value == "毛利率"
        assert ws.cell(row=1, column=5).value == "备注"


@pytest.mark.skipif(
    importlib.util.find_spec("openpyxl") is None,
    reason="openpyxl only installed in the mcp container",
)
def test_chart_builder_quotes_sheet_names_with_spaces():
    """Regression: a sheet name containing a space (e.g. 'Q3 销售') must be
    wrapped in single quotes inside the structured-reference range string,
    otherwise openpyxl parses 'Q3 销售!B1:B3' as the unknown sheet 'Q3' +
    stray garbage and raises ValueError.
    """
    from xlsx_engine._handle import use_workdir
    from xlsx_engine import builder, chart_builder

    with tempfile.TemporaryDirectory() as tmp:
        with use_workdir(tmp):
            builder.create_workbook(
                filename="in.xlsx",
                sheets=[{
                    "name": "Q3 销售",
                    "headers": ["地区", "收入"],
                    "rows": [["华东", 125], ["华南", 98], ["华北", 72]],
                }],
            )
            chart_builder.add_chart(
                input_filename="in.xlsx",
                output_filename="out.xlsx",
                sheet="Q3 销售",
                chart_type="bar",
                data_range="B1:B4",
                categories_range="A2:A4",
                title="Q3 各地区收入对比",
                anchor="G2",
            )

    # The unit-level helper covers the quoting grammar; here we just verify
    # the end-to-end add_chart didn't raise on a sheet with a space.


def test_chart_builder_quote_sheet_ref_helper():
    """Quoting rules: identifier-safe names stay bare; everything else gets
    single-quoted with embedded quotes doubled (OOXML grammar)."""
    from xlsx_engine.chart_builder import _quote_sheet_ref

    assert _quote_sheet_ref("Sheet1") == "Sheet1"
    assert _quote_sheet_ref("Data") == "Data"
    assert _quote_sheet_ref("Q3 销售") == "'Q3 销售'"
    assert _quote_sheet_ref("3rd Quarter") == "'3rd Quarter'"
    assert _quote_sheet_ref("foo'bar") == "'foo''bar'"


@pytest.mark.skipif(
    importlib.util.find_spec("openpyxl") is None,
    reason="openpyxl only installed in the mcp container",
)
def test_chart_builder_embeds_chart():
    from xlsx_engine._handle import use_workdir
    from xlsx_engine import chart_builder, model_builder

    with tempfile.TemporaryDirectory() as tmp:
        with use_workdir(tmp):
            model_builder.build_model(
                spec={"sheets": [{"name": "Data", "rows": [
                    {"role": "header", "cells": ["Q", "V"]},
                    {"cells": ["Q1", 10]}, {"cells": ["Q2", 22]},
                    {"cells": ["Q3", 35]}]}]},
                output_filename="d.xlsx")
            chart_builder.add_chart(
                input_filename="d.xlsx", output_filename="dc.xlsx",
                sheet="Data", chart_type="bar",
                data_range="B1:B4", categories_range="A2:A4", title="Q")
        out = Path(tmp) / "dc.xlsx"
        assert out.is_file()
        with zipfile.ZipFile(out) as z:
            assert any("chart" in n for n in z.namelist())
