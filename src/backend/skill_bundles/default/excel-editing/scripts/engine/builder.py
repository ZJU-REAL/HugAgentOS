"""Excel workbook creation primitives (openpyxl).

Inspired by ``agent_skills/skills/minimax-xlsx/scripts/xlsx_cli.py``'s "Workbook spec"
pattern — a single JSON structure describes sheets, headers, columns, and rows.
We keep the spec shallower than xlsx_cli's (which supports role-based styling,
formulas, and multi-sheet financial models) to keep the schema obvious to LLMs.
The full Formula-First / role-styling power can be reintroduced as separate
tools (excel_apply_formula, excel_apply_role_styling) without changing this
foundational builder.
"""
from __future__ import annotations

from typing import Any, Iterable

from ._handle import input_path, output_path


def _set_column_widths(ws, widths: list[float] | None) -> None:
    """Apply column widths (in Excel character units) by index."""
    if not widths:
        return
    from openpyxl.utils import get_column_letter

    for idx, width in enumerate(widths, start=1):
        if width and width > 0:
            ws.column_dimensions[get_column_letter(idx)].width = float(width)


def _write_rows(ws, rows: Iterable[list[Any]], *, start_row: int = 1) -> int:
    """Write rows starting at ``start_row``; return the next free row index."""
    row_idx = start_row
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1
    return row_idx


def _apply_header_style(ws, header_row: int, n_cols: int) -> None:
    """Bold + filled header row (matches the look of report_export_mcp tables)."""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def create_workbook(
    *,
    filename: str,
    sheets: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Create a new .xlsx with one or more sheets.

    Args:
        filename: output filename (must end with ``.xlsx``)
        sheets: list of sheet specs. Each:
            ``{
              "name": str,                    # sheet tab name (default Sheet1, Sheet2, ...)
              "headers": list[str]?,          # if provided, written as styled row 1
              "rows": list[list[Any]]?,       # data rows (after headers, if any)
              "column_widths": list[float]?,  # optional per-column widths
              "freeze_header": bool?,         # if True and headers given, freeze row 1
            }``
            If sheets is None or empty, an empty default sheet is created.

    Returns:
        ``{"output_filename", "sheet_names", "row_counts": {sheet_name: int}}``
    """
    from openpyxl import Workbook

    wb = Workbook()
    # openpyxl creates a default 'Sheet'; we'll repurpose or remove it
    default_ws = wb.active
    default_used = False

    sheet_specs = sheets or [{"name": "Sheet1"}]
    sheet_names: list[str] = []
    row_counts: dict[str, int] = {}

    for i, spec in enumerate(sheet_specs):
        name = (spec.get("name") or f"Sheet{i+1}").strip() or f"Sheet{i+1}"
        if i == 0:
            default_ws.title = name
            ws = default_ws
            default_used = True
        else:
            ws = wb.create_sheet(title=name)

        headers = spec.get("headers") or []
        rows = spec.get("rows") or []

        next_row = 1
        if headers:
            next_row = _write_rows(ws, [headers], start_row=1)
            _apply_header_style(ws, header_row=1, n_cols=len(headers))
            if spec.get("freeze_header"):
                ws.freeze_panes = "A2"
        next_row = _write_rows(ws, rows, start_row=next_row)

        _set_column_widths(ws, spec.get("column_widths"))

        sheet_names.append(name)
        row_counts[name] = (next_row - 1) - (0 if not headers else 0)

    if not default_used and default_ws.title not in sheet_names:
        # No sheet ever claimed the default; remove it
        wb.remove(default_ws)

    out = output_path(filename)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    return {
        "output_filename": filename,
        "sheet_names": sheet_names,
        "row_counts": row_counts,
        "size_bytes": out.stat().st_size,
    }
