"""Read operations on .xlsx — summary, sheet data extraction.

Inspired by ``agent_skills/skills/minimax-xlsx/scripts/xlsx_reader.py`` —
the original returns pandas DataFrames; here we return plain dict/list so
the result serializes through ``__OFFICE_MCP_RESULT_v1__:`` cleanly without
requiring pandas in either the result-shape or the sandbox kernel state
(pandas is installed but pulling DataFrames through JSON is wasteful).
"""
from __future__ import annotations

from typing import Any

from ._handle import input_path


_DEFAULT_SAMPLE_ROWS = 5


def get_summary(
    *,
    input_filename: str,
    sample_rows: int = _DEFAULT_SAMPLE_ROWS,
) -> dict[str, Any]:
    """Return high-level workbook info: sheet list, dimensions, sample rows per sheet.

    Args:
        input_filename: source .xlsx in cwd
        sample_rows: number of rows from each sheet to include in ``sample`` (max 50)

    Returns:
        ``{"sheet_names": [...],
            "sheets": [{"name", "max_row", "max_column",
                        "headers": [...], "sample": [[...], ...]}, ...]}``
    """
    from openpyxl import load_workbook

    sample_rows = max(0, min(int(sample_rows), 50))
    wb = load_workbook(filename=str(input_path(input_filename)), data_only=True, read_only=True)

    sheets: list[dict[str, Any]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Pull header row + sample rows. ``ws.iter_rows`` is the read_only-safe path.
        rows = list(ws.iter_rows(min_row=1, max_row=min(max_row, sample_rows + 1), values_only=True))
        headers = list(rows[0]) if rows else []
        sample = [list(r) for r in rows[1 : sample_rows + 1]] if rows else []

        sheets.append(
            {
                "name": name,
                "max_row": max_row,
                "max_column": max_col,
                "headers": headers,
                "sample": sample,
            }
        )

    wb.close()
    return {
        "sheet_names": list(wb.sheetnames),
        "sheets": sheets,
    }


def get_sheet_data(
    *,
    input_filename: str,
    sheet: str,
    cell_range: str | None = None,
    max_rows: int = 1000,
) -> dict[str, Any]:
    """Return cell values for a sheet (optionally limited to a range like ``"A1:D20"``).

    Args:
        input_filename: source .xlsx in cwd
        sheet: target sheet name
        cell_range: openpyxl-style range ``"A1:D20"``; None = entire used area
        max_rows: hard cap on rows returned (default 1000) to bound LLM context

    Returns:
        ``{"sheet", "range", "row_count", "column_count", "rows": [[...], ...]}``
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(input_path(input_filename)), data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"sheet {sheet!r} not found; available: {wb.sheetnames}")

    ws = wb[sheet]

    if cell_range:
        # Use openpyxl's range slicing
        try:
            cell_block = ws[cell_range]
        except Exception as e:
            wb.close()
            raise ValueError(f"invalid cell_range {cell_range!r}: {e}") from e
        # ws[range] returns a tuple of tuples of Cell objects (or a single tuple for single row/col)
        if hasattr(cell_block, "__iter__"):
            rows_iter = cell_block
        else:
            rows_iter = (cell_block,)
        rows = []
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                break
            if hasattr(row, "value"):
                # Single cell row
                rows.append([row.value])
            else:
                rows.append([cell.value for cell in row])
    else:
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(list(row))

    wb.close()
    return {
        "sheet": sheet,
        "range": cell_range,
        "row_count": len(rows),
        "column_count": len(rows[0]) if rows else 0,
        "rows": rows,
    }
