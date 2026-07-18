"""Excel cell + sheet editing (openpyxl)."""
from __future__ import annotations

from typing import Any

from ._handle import input_path, output_path


def set_cells(
    *,
    input_filename: str,
    output_filename: str,
    sheet: str,
    cells: list[dict[str, Any]],
) -> dict[str, Any]:
    """Set cell values / formulas in an existing workbook.

    Args:
        input_filename: source xlsx
        output_filename: destination xlsx
        sheet: sheet name to operate on
        cells: list of ``{addr, value, formula?}``::

            [{"addr": "B2", "value": 100},
             {"addr": "C2", "formula": "=SUM(B2:B10)"}]

            ``formula`` (when set) takes precedence over ``value``. Formula
            strings may include or omit the leading ``=``; both work.

    Returns:
        ``{"output_filename", "sheet", "cells_written": N}``
    """
    from openpyxl import load_workbook

    if not cells:
        raise ValueError("'cells' must be a non-empty list")

    wb = load_workbook(filename=str(input_path(input_filename)))
    if sheet not in wb.sheetnames:
        raise ValueError(
            f"sheet {sheet!r} not found; available: {wb.sheetnames}"
        )
    ws = wb[sheet]

    written = 0
    for spec in cells:
        addr = spec.get("addr")
        if not addr:
            raise ValueError(f"cell entry missing 'addr': {spec}")
        if "formula" in spec and spec["formula"]:
            f = str(spec["formula"])
            ws[addr] = f if f.startswith("=") else "=" + f
        elif "value" in spec:
            ws[addr] = spec["value"]
        else:
            raise ValueError(f"cell entry must have 'value' or 'formula': {spec}")
        written += 1

    out = output_path(output_filename)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    return {
        "output_filename": output_filename,
        "sheet": sheet,
        "cells_written": written,
    }


def add_sheet(
    *,
    input_filename: str,
    output_filename: str,
    sheet_name: str,
    after: str | None = None,
    headers: list[str] | None = None,
) -> dict[str, Any]:
    """Append a new sheet to an existing workbook.

    Args:
        input_filename: source xlsx
        output_filename: destination xlsx
        sheet_name: tab name for the new sheet (must not already exist)
        after: optional name of an existing sheet; new sheet is inserted right
               after it. If None, the new sheet is appended at the end.
        headers: optional row 1 to write (auto-styled when present)

    Returns:
        ``{"output_filename", "sheet_name", "sheet_index", "sheet_count"}``
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    if not sheet_name or not sheet_name.strip():
        raise ValueError("'sheet_name' must be non-empty")

    wb = load_workbook(filename=str(input_path(input_filename)))
    if sheet_name in wb.sheetnames:
        raise ValueError(f"sheet {sheet_name!r} already exists; use a different name")

    if after is not None:
        if after not in wb.sheetnames:
            raise ValueError(f"after-sheet {after!r} not found; available: {wb.sheetnames}")
        idx = wb.sheetnames.index(after) + 1
    else:
        idx = len(wb.sheetnames)

    ws = wb.create_sheet(title=sheet_name, index=idx)

    if headers:
        for col_idx, value in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=value)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    out = output_path(output_filename)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    return {
        "output_filename": output_filename,
        "sheet_name": sheet_name,
        "sheet_index": idx,
        "sheet_count": len(wb.sheetnames),
    }
