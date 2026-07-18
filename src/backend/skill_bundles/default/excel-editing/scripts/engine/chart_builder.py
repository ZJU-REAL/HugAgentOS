"""Native Excel chart insertion (openpyxl) — net-new (not in minimax-xlsx).

NOTE: unlike ``patch_editor`` / ``model_builder`` (byte-preserving), this uses
an openpyxl load→save round-trip. That can drop VBA / pivot tables / advanced
conditional formatting. Use it on workbooks you created here or where a plain
data workbook is acceptable.

Public: add_chart(input_filename, output_filename, sheet, chart_type,
                   data_range, ...) -> dict
"""
from __future__ import annotations

import re
from typing import Any

from ._handle import input_path, output_path

_CHART_TYPES = {"bar", "line", "pie"}

# Sheet names that can appear unquoted in a structured Excel reference.
# Anything containing a space, punctuation, or starting with a digit must be
# wrapped in single quotes (OOXML reference grammar). openpyxl's Reference
# parser respects this — without the quotes you get `Q3 销售!B1:B6` which is
# parsed as the unknown sheet "Q3" + stray garbage and fails the lookup.
_UNQUOTED_SHEET_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _quote_sheet_ref(sheet: str) -> str:
    """Return the sheet name as it should appear in a `<sheet>!<range>` ref."""
    if _UNQUOTED_SHEET_RE.match(sheet):
        return sheet
    # Per OOXML, embedded single quotes are doubled inside the quoted form.
    return "'" + sheet.replace("'", "''") + "'"


def add_chart(
    *,
    input_filename: str,
    output_filename: str,
    sheet: str,
    chart_type: str,
    data_range: str,
    categories_range: str | None = None,
    title: str | None = None,
    x_title: str | None = None,
    y_title: str | None = None,
    anchor: str = "H2",
) -> dict[str, Any]:
    """Add a bar / line / pie chart to a sheet.

    Args:
        sheet:            sheet that holds the data AND receives the chart
        chart_type:       "bar" | "line" | "pie"
        data_range:       values range, e.g. "B1:B10" (include a header row for
                          the series title)
        categories_range: optional axis-label range, e.g. "A2:A10"
        title/x_title/y_title: optional labels
        anchor:           top-left cell where the chart is placed (default H2)
    """
    if chart_type not in _CHART_TYPES:
        raise ValueError(
            f"chart_type {chart_type!r} not supported; choose from {sorted(_CHART_TYPES)}"
        )

    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    wb = load_workbook(filename=str(input_path(input_filename)))
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet {sheet!r} not found; available: {wb.sheetnames}")
    ws = wb[sheet]

    if chart_type == "bar":
        chart = BarChart()
        chart.type = "col"
    elif chart_type == "line":
        chart = LineChart()
    else:
        chart = PieChart()

    if title:
        chart.title = title
    if chart_type != "pie":
        if x_title:
            chart.x_axis.title = x_title
        if y_title:
            chart.y_axis.title = y_title

    sheet_ref = _quote_sheet_ref(sheet)
    data = Reference(ws, range_string=f"{sheet_ref}!{data_range}")
    chart.add_data(data, titles_from_data=True)

    if categories_range:
        cats = Reference(ws, range_string=f"{sheet_ref}!{categories_range}")
        chart.set_categories(cats)

    ws.add_chart(chart, anchor)

    out = output_path(output_filename)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    return {
        "output_filename": output_filename,
        "sheet": sheet,
        "chart_type": chart_type,
        "anchor": anchor,
        "size_bytes": out.stat().st_size,
    }
